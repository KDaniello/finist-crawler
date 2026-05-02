import csv
import json
import logging
import multiprocessing
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)
__all__ = ["DataWriter", "SessionManager"]


class SessionManager:
    def __init__(self, base_dir: Path) -> None:
        self.base_dir = base_dir

    def create_session(self) -> str:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S-%f")
        session_id = f"session_{timestamp}"
        session_path = self.base_dir / session_id

        try:
            session_path.mkdir(parents=True, exist_ok=True)
            logger.info(f"Создана новая сессия: {session_id}")
            return session_id
        except OSError as e:
            logger.critical(f"Ошибка создания папки сессии: {e}")
            raise RuntimeError(f"Не удалось инициализировать сессию: {e}") from e

    def list_sessions(self) -> list[str]:
        if not self.base_dir.exists():
            return []
        sessions = [
            d.name for d in self.base_dir.iterdir() if d.is_dir() and d.name.startswith("session_")
        ]
        return sorted(sessions, reverse=True)


# Поля с длинным текстом — в Excel будут широкими с переносом
_WIDE_TEXT_FIELDS = {"text", "excerpt", "description", "alt_headline", "body", "content"}

# Максимальная ширина колонки в Excel (в символах)
_MAX_COL_WIDTH = 80
_MIN_COL_WIDTH = 10


def _clean_for_csv(value: Any) -> str:
    """
    Подготавливает значение для записи в CSV.
    Заменяет переносы строк на пробел — иначе Excel/LibreOffice
    воспринимает их как новые строки таблицы.
    """
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    text = str(value)
    # Нормализуем все виды переносов строк
    text = text.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    # Убираем задвоенные пробелы
    while "  " in text:
        text = text.replace("  ", " ")
    return text.strip()


def _clean_for_excel(value: Any) -> Any:
    """
    Подготавливает значение для записи в Excel.
    В XLSX переносы строк внутри ячейки — это норма,
    поэтому оставляем текст как есть.
    Только dict/list конвертируем в JSON-строку.
    """
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value)


class DataWriter:
    def __init__(
        self,
        base_dir: Path,
        session_id: str,
        source: str,
        lock: multiprocessing.synchronize.Lock,
    ) -> None:
        self.session_id = session_id
        self.source = source
        self.base_dir = base_dir
        self.source_dir = self.base_dir / session_id / source
        self.jsonl_path = self.source_dir / f"{source}.jsonl"
        self._lock = lock

    def save_batch(self, data: Iterable[dict[str, Any]]) -> None:
        records = list(data)
        if not records:
            return

        try:
            self.source_dir.mkdir(parents=True, exist_ok=True)

            with self._lock, open(self.jsonl_path, "a", encoding="utf-8") as f:
                for record in records:
                    f.write(json.dumps(record, ensure_ascii=False) + "\n")
                f.flush()

            logger.debug(f"Сохранено {len(records)} записей в {self.jsonl_path.name}")
        except Exception as e:
            logger.error(f"Ошибка записи батча [{self.source}]: {e}", exc_info=True)
            raise

    def export(self, fmt: str = "xlsx") -> Path | None:
        if not self.jsonl_path.exists():
            logger.warning(f"Нет данных для экспорта [{self.source}] в сессии {self.session_id}")
            return None

        output_path = self.source_dir / f"{self.source}_export.{fmt}"
        try:
            if fmt == "csv":
                self._export_to_csv(output_path)
            elif fmt == "xlsx":
                self._export_to_excel(output_path)
            else:
                raise ValueError(f"Неизвестный формат экспорта: {fmt}")
            logger.info(f"Успешный экспорт: {output_path.name}")
            return output_path
        except Exception as e:
            logger.error(f"Ошибка экспорта {self.source} в {fmt}: {e}", exc_info=True)
            raise

    def _export_to_csv(self, output_path: Path) -> None:
        """
        CSV-экспорт. Переносы строк в тексте заменяются пробелами —
        это необходимо для корректного отображения в Excel/LibreOffice.
        Для полного текста со структурой используйте XLSX.
        """
        with open(self.jsonl_path, encoding="utf-8") as fin:
            with open(output_path, "w", encoding="utf-8-sig", newline="") as fout:
                writer = None
                for line in fin:
                    if not line.strip():
                        continue
                    record = json.loads(line)
                    if writer is None:
                        writer = csv.DictWriter(
                            fout,
                            fieldnames=list(record.keys()),
                            delimiter=";",
                        )
                        writer.writeheader()

                    # Чистим все значения от переносов строк
                    safe_record = {k: _clean_for_csv(record.get(k, "")) for k in writer.fieldnames}
                    writer.writerow(safe_record)

    def _export_to_excel(self, output_path: Path) -> None:
        """
        XLSX-экспорт с форматированием:
        - Заголовки: жирный шрифт, серый фон
        - Текстовые поля: перенос по словам (wrap_text)
        - Автоподбор ширины колонок (с ограничением)
        - Заморозка первой строки (freeze_panes)
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.source[:31]

        # Стили заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers_written = False
        keys: list[str] = []
        col_widths: dict[int, int] = {}  # col_idx -> max_width

        with open(self.jsonl_path, encoding="utf-8") as fin:
            for row_idx, line in enumerate(fin, start=2):  # строки данных с 2-й
                if not line.strip():
                    continue
                record = json.loads(line)

                # Пишем заголовки один раз
                if not headers_written:
                    keys = list(record.keys())
                    for col_idx, key in enumerate(keys, start=1):
                        cell = ws.cell(row=1, column=col_idx, value=key)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        col_widths[col_idx] = max(len(key), _MIN_COL_WIDTH)
                    headers_written = True

                # Пишем данные
                for col_idx, key in enumerate(keys, start=1):
                    raw_val = record.get(key, "")
                    val = _clean_for_excel(raw_val)
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)

                    is_text_field = key in _WIDE_TEXT_FIELDS
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=is_text_field,
                    )

                    # Считаем максимальную ширину для автоподбора
                    val_str = str(val) if val is not None else ""
                    # Для многострочных берём длину самой длинной строки
                    max_line = max(
                        (len(line) for line in val_str.split("\n")),
                        default=0,
                    )
                    current_width = min(max_line + 2, _MAX_COL_WIDTH)
                    col_widths[col_idx] = max(col_widths.get(col_idx, 0), current_width)

        if not headers_written:
            logger.warning(f"Нет данных для Excel экспорта [{self.source}]")
            return

        # Применяем ширину колонок
        for col_idx, width in col_widths.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(width, _MIN_COL_WIDTH)

        # Высота строки заголовков
        ws.row_dimensions[1].height = 30

        # Заморозка первой строки — при прокрутке заголовки остаются видны
        ws.freeze_panes = "A2"

        # Автофильтр на заголовки
        ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
